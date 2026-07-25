# -*- coding: utf-8 -*-
"""Emite src/data/catalog.ts con el catálogo (edificio + puente): materiales,
componentes->elementos(materiales,zonas), deterioros(materiales,gravedad) y causas.
Puente: se lee de Gravedades_Base.xlsx y se ENRIQUECE con materiales + anclajes."""
import json
import openpyxl

MATERIALS = [
    "Hormigón armado", "Hormigón pretensado/postensado", "Acero", "Madera",
    "Albañilería confinada", "Albañilería armada", "Albañilería simple",
    "Mixto acero-hormigón", "Elastomérico", "Otro",
]
HA, HP, AC, MA = "Hormigón armado", "Hormigón pretensado/postensado", "Acero", "Madera"
ALBC, ALBA, MX = "Albañilería confinada", "Albañilería armada", "Mixto acero-hormigón"


def zl(s):
    return [z.strip() for z in s.split(";") if z.strip()]


# ───────────── EDIFICACIÓN ─────────────
edif_components = {
    "Fundaciones": [
        ("Zapata aislada", [HA], "Cara superior; Pedestal; Perímetro; Otro"),
        ("Zapata corrida", [HA], "Cara superior; Cara lateral; Otro"),
        ("Viga de fundación", [HA], "Apoyo; Centro de vano; Nudo; Otro"),
        ("Losa de fundación (radier)", [HA], "Centro de paño; Borde; Junta; Otro"),
        ("Pilote", [HA, AC, MA], "Cabeza; Fuste; Punta; Anclaje/empotramiento; Otro"),
        ("Cabezal de pilotes", [HA], "Cara superior; Encuentro con pilote; Otro"),
        ("Sobrecimiento", [HA, ALBC], "Base; Coronación; Otro"),
        ("Muro de subterráneo", [HA], "Base; Coronación; Paño central; Junta; Otro"),
    ],
    "Muros estructurales": [
        ("Muro estructural / de corte", [HA], "Base; Coronación; Paño central; Encuentro/esquina; Vano/abertura; Otro"),
        ("Viga de acople (dintel de acople)", [HA], "Empotramiento en muro; Tramo/centro; Armadura diagonal; Otro"),
        ("Machón", [HA, ALBC], "Base; Cuerpo; Encuentro; Otro"),
        ("Muro de albañilería confinada", [ALBC], "Paño central; Encuentro; Base; Bajo cadena/pilar; Vano; Otro"),
        ("Muro de albañilería armada", [ALBA], "Paño central; Encuentro; Base; Vano; Otro"),
        ("Muro de contención", [HA], "Coronación; Fuste/pantalla; Base/empotramiento; Puntera; Talón; Junta; Drenaje/barbacana; Otro"),
    ],
    "Pilares y columnas": [
        ("Columna / Pilar", [HA, AC, MA, MX], "Base; Anclaje / placa base; Tercio central; Cabeza/capitel; Nudo; Empalme; Otro"),
    ],
    "Vigas y cadenas": [
        ("Viga", [HA, HP, AC, MA, MX], "Apoyo/extremo; Centro de vano; Fondo de viga; Nudo; Anclaje de pretensado; Anclaje de conexión; Cara lateral; Otro"),
        ("Cadena", [HA], "Tramo; Encuentro; Otro"),
        ("Dintel", [HA, AC, ALBC], "Centro; Apoyo; Otro"),
    ],
    "Losas y entrepisos": [
        ("Losa", [HA, HP], "Centro de paño; Borde/perímetro; Cara inferior; Cara superior; Junta; Perforación/paso de instalaciones; Zona de punzonamiento (sobre columna); Otro"),
        ("Losa colaborante", [MX], "Centro de paño; Apoyo; Conexión de corte; Otro"),
        ("Vigueta", [HA, HP, AC, MA], "Apoyo; Centro de vano; Fondo; Otro"),
        ("Voladizo / balcón", [HA], "Empotramiento; Extremo; Cara inferior; Otro"),
    ],
    "Nudos y conexiones": [
        ("Nudo viga-columna", [HA, AC], "Núcleo del nudo; Extremo de viga; Extremo de columna; Otro"),
        ("Conexión soldada", [AC], "Cordón de soldadura; Placa; Otro"),
        ("Conexión apernada", [AC], "Perno; Placa; Otro"),
        ("Anclaje", [AC, HA], "Placa base; Perno de anclaje; Anclaje de pretensado; Anclaje químico; Anclaje mecánico; Otro"),
        ("Empalme de armadura", [HA], "Traslapo; Manguito; Otro"),
        ("Apoyo / junta de dilatación", [HA, AC, "Elastomérico"], "Aparato de apoyo; Junta; Otro"),
    ],
    "Escaleras": [
        ("Losa de escalera", [HA], "Tramo; Descanso; Apoyo; Otro"),
        ("Peldañeado", [HA, MA, AC], "Peldaño; Nariz; Otro"),
        ("Zanca / viga de escalera", [HA, AC, MA], "Apoyo; Tramo; Anclaje; Otro"),
        ("Rampa", [HA], "Tramo; Apoyo; Otro"),
    ],
    "Techumbre estructural": [
        ("Cercha", [AC, MA], "Cordón superior; Cordón inferior; Diagonal; Montante; Nudo; Anclaje/apoyo; Otro"),
        ("Tijeral", [MA, AC], "Cordón superior; Cordón inferior; Diagonal; Pendolón; Nudo; Apoyo/anclaje; Otro"),
        ("Viga de techo", [AC, MA, HA], "Apoyo; Centro de vano; Anclaje; Otro"),
        ("Costanera / correa", [AC, MA], "Apoyo; Tramo; Conexión; Otro"),
        ("Entramado / diafragma de techumbre", [MA, AC], "Paño; Arriostramiento; Encuentro/anclaje; Otro"),
        ("Losa de cubierta", [HA], "Centro de paño; Borde; Junta; Otro"),
    ],
    "Tabiquería": [
        ("Tabique", [ALBC, "Otro"], "Paño central; Encuentro; Base; Vano; Otro"),
        ("Antepecho / parapeto", [HA, ALBC], "Coronación; Base; Otro"),
    ],
}
TODOS = "todos"
edif_damages = {
    "Fisuras": [HA, "Albañilería confinada", MA],
    "Grietas": [HA, "Albañilería confinada"],
    "Fisuras en mapa o retícula": [HA],
    "Descascaramiento sin armadura a la vista": [HA],
    "Descascaramiento con armadura a la vista": [HA],
    "Armadura a la vista": [HA],
    "Corrosión de armaduras/acero": [HA, AC],
    "Mancha de óxido": [HA, AC],
    "Coqueras / nidos de grava": [HA],
    "Humedades / filtraciones": [TODOS],
    "Eflorescencias": [HA, "Albañilería confinada"],
    "Deformación / flecha excesiva": [TODOS],
    "Falta de alineación / desplome": [TODOS],
    "Disgregación / pulverización": [HA, "Albañilería confinada"],
    "Aplastamiento (albañilería)": ["Albañilería confinada", HA],
    "Pandeo local": [AC],
    "Pérdida de mortero de junta": ["Albañilería confinada"],
    "Pérdida de tratamiento protector": [AC, MA],
    "Pudrición / xilófagos": [MA],
    "Deterioro de soldadura / conexión": [AC],
    "Pérdida / aflojamiento de pernos y anclajes": [AC, HA],
    "Alteración superficial": [TODOS],
    "Vegetación / biológico": [TODOS],
}
edif_severity = {
    "Fisuras": ["abertura < 0,2 mm, sin incidencia estructural", "0,2-0,4 mm o incidencia estructural leve", "0,4-1 mm con incidencia estructural", "> 1 mm, evolutiva o generalizada"],
    "Grietas": ["~1 mm, pasante, sin incidencia estructural aparente", "1-2 mm con incidencia estructural", "2-5 mm", "> 5 mm, evolutiva; compromete estabilidad"],
    "Fisuras en mapa o retícula": ["fisuración superficial fina, sin desprendimientos", "retícula marcada con inicio de disgregación superficial", "fisuración con desprendimiento de árido y avance en profundidad", "generalizada, con pérdida de material y afectación de armadura"],
    "Descascaramiento sin armadura a la vista": ["desprendimiento puntual del recubrimiento, sin llegar a la armadura", "descascaramiento localizado, recubrimiento reducido", "descascaramiento extenso, armadura a punto de quedar expuesta", "pérdida generalizada del recubrimiento en la zona"],
    "Descascaramiento con armadura a la vista": ["sin pérdida de sección de armadura", "pérdida leve (<10%)", "pérdida apreciable (10-50%)", "pérdida > 50% o barras seccionadas"],
    "Armadura a la vista": ["armadura expuesta sin corrosión, puntual", "armadura expuesta con oxidación superficial", "expuesta con corrosión y pérdida de sección leve", "expuesta y corroída, con pérdida de sección significativa o barras seccionadas"],
    "Corrosión de armaduras/acero": ["oxidación superficial, <10%, sin pérdida", "pérdida leve (10-25%)", "pérdida apreciable (25-50%)", "pérdida > 50% o seccionamiento"],
    "Mancha de óxido": ["mancha superficial aislada, sin fisura asociada", "manchas recurrentes, posible corrosión incipiente bajo recubrimiento", "manchas con fisura o abombamiento del recubrimiento", "generalizadas, con descascaramiento y armadura corroída"],
    "Coqueras / nidos de grava": ["coquera superficial pequeña, sin armadura expuesta", "coquera localizada con recubrimiento reducido", "coquera profunda con armadura expuesta", "nidos extensos que comprometen sección o anclaje de armadura"],
    "Humedades / filtraciones": ["mancha superficial, sin daño", "humedad persistente, inicio de deterioro", "filtración activa con daño", "filtración severa con corrosión/pérdida"],
    "Eflorescencias": ["depósito salino leve, superficie seca", "eflorescencia con humedad asociada", "persistente, con inicio de disgregación/lixiviación", "lixiviación activa (lavado/estalactitas) con deterioro del material"],
    "Deformación / flecha excesiva": ["perceptible, < L/500", "L/500-L/250", "L/250-L/150", "> L/150 o creciente"],
    "Falta de alineación / desplome": ["desviación perceptible dentro de tolerancias, sin daño asociado", "desplome apreciable, sin fisuración estructural", "desplome con fisuración asociada / fuera de tolerancia", "desplome severo o creciente que compromete la estabilidad"],
    "Disgregación / pulverización": ["disgregación superficial leve, material pulverulento puntual", "pérdida de mortero/árido superficial", "disgregación en profundidad con reducción de sección", "pérdida de material generalizada con armadura expuesta o comprometida"],
    "Aplastamiento (albañilería)": ["fisuras finas por compresión, sin desprendimiento", "fisuración vertical marcada, inicio de desconche de unidades", "aplastamiento local con desprendimiento de piezas", "generalizado, con pérdida de capacidad portante"],
    "Pandeo local": ["abolladura leve de ala/alma, sin plastificación", "pandeo local apreciable, sin pérdida de capacidad evidente", "pandeo con plastificación local", "pandeo severo con pérdida de capacidad o colapso local inminente"],
    "Pérdida de mortero de junta": ["pérdida superficial de mortero (< 1 cm)", "juntas erosionadas, profundidad moderada", "pérdida en profundidad, holgura entre piezas", "juntas vacías generalizadas, inestabilidad de piezas"],
    "Pérdida de tratamiento protector": ["deterioro puntual de pintura/galvanizado, sin corrosión", "pérdida localizada con oxidación superficial incipiente", "pérdida extensa con corrosión activa", "ausencia total de protección con corrosión avanzada"],
    "Pudrición / xilófagos": ["ataque superficial o indicios (galerías), sin pérdida de sección", "pudrición o ataque localizado con pérdida de sección leve", "pudrición en profundidad con pérdida de sección apreciable", "pérdida de sección severa; sin capacidad resistente fiable"],
    "Deterioro de soldadura / conexión": ["defecto superficial (salpicadura, mordedura leve), sin fisura", "corrosión o fisura fina en el cordón, sin pérdida de capacidad", "fisura en soldadura o pérdida de garganta apreciable", "fisura pasante o rotura de la conexión"],
    "Pérdida / aflojamiento de pernos y anclajes": ["pernos con oxidación superficial, apriete correcto", "aflojamiento leve o falta puntual de pernos no críticos", "varios pernos flojos/faltantes o corrosión con pérdida de sección", "pérdida generalizada o anclaje sin capacidad; conexión comprometida"],
    "Alteración superficial": ["suciedad, pátina o decoloración, sin daño del material", "erosión/abrasión superficial leve", "alteración con pérdida de material superficial", "alteración profunda que reduce sección o recubrimiento"],
    "Vegetación / biológico": ["presencia de musgo/líquenes, sin daño", "vegetación menor con retención de humedad", "raíces o vegetación con fisuración/disgregación asociada", "vegetación arraigada que abre juntas o desplaza elementos"],
}
edif_causes = {
    "Fisuras": ["retracción del hormigón", "dilatación/contracción térmica", "curado inadecuado", "asiento diferencial", "sobrecarga", "deficiente ejecución", "causa desconocida"],
    "Grietas": ["acción sísmica", "asiento diferencial", "sobrecarga gravitacional", "esfuerzos (corte)", "esfuerzos (flexión)", "corrosión de armaduras", "empuje de terreno", "deficiente diseño", "causa desconocida"],
    "Fisuras en mapa o retícula": ["retracción", "ataque químico", "reacción árido-álcali", "corrosión de armaduras", "ciclos hielo-deshielo", "deficiente ejecución"],
    "Descascaramiento sin armadura a la vista": ["golpe o impacto", "ciclos hielo-deshielo", "ataque químico", "deficiente ejecución", "causa desconocida"],
    "Descascaramiento con armadura a la vista": ["corrosión de armaduras", "carbonatación", "ataque de cloruros", "escasez de recubrimiento", "deficiente ejecución"],
    "Armadura a la vista": ["escasez de recubrimiento", "corrosión de armaduras", "desconchón previo", "deficiente ejecución"],
    "Corrosión de armaduras/acero": ["carbonatación", "ataque de cloruros", "humedad/filtración", "escasez de recubrimiento", "ambiente agresivo/marino", "falta de mantención"],
    "Mancha de óxido": ["corrosión de armaduras", "humedad", "escorrentía superficial", "falta de limpieza del encofrado"],
    "Coqueras / nidos de grava": ["deficiente ejecución (vibrado)", "deficiente dosificación", "causa desconocida"],
    "Humedades / filtraciones": ["impermeabilización defectuosa", "rotura de conducción", "capilaridad", "condensación", "mal drenaje", "falta de mantención"],
    "Eflorescencias": ["humedad/filtración", "capilaridad", "impermeabilización defectuosa", "sales del material"],
    "Deformación / flecha excesiva": ["sobrecarga", "infradimensionamiento", "fluencia (creep) del hormigón", "cambio de uso", "esfuerzos (flexión)", "pandeo"],
    "Falta de alineación / desplome": ["asiento diferencial", "acción sísmica", "empuje de terreno", "deficiente ejecución"],
    "Disgregación / pulverización": ["ataque químico", "ciclos hielo-deshielo", "carbonatación avanzada", "curado inadecuado", "calidad deficiente del hormigón"],
    "Aplastamiento (albañilería)": ["sobrecarga", "acción sísmica", "infradimensionamiento", "concentración de esfuerzos"],
    "Pandeo local": ["sobrecarga (compresión)", "esbeltez excesiva", "arriostramiento insuficiente", "infradimensionamiento", "impacto"],
    "Pérdida de mortero de junta": ["envejecimiento", "humedad", "ataque químico", "deficiente ejecución", "acción sísmica"],
    "Pérdida de tratamiento protector": ["envejecimiento", "acción climática", "falta de mantención", "corrosión", "ataque químico"],
    "Pudrición / xilófagos": ["humedad", "hongos", "insectos xilófagos", "falta de protección/mantención"],
    "Deterioro de soldadura / conexión": ["fatiga", "corrosión", "deficiente ejecución", "esfuerzos cíclicos", "sobrecarga"],
    "Pérdida / aflojamiento de pernos y anclajes": ["corrosión", "aflojamiento por vibración", "deficiente ejecución", "esfuerzos cíclicos", "sobrecarga", "acción sísmica"],
    "Alteración superficial": ["acción climática", "envejecimiento", "ataque químico", "abrasión", "pérdida de tratamiento"],
    "Vegetación / biológico": ["humedad", "sedimentación orgánica", "falta de mantención"],
}


def build_edificio():
    comps = []
    for comp, elems in edif_components.items():
        comps.append({"component": comp, "elements": [
            {"element": n, "materials": m, "zones": zl(z)} for (n, m, z) in elems]})
    damages = []
    for d, mats in edif_damages.items():
        sev = edif_severity.get(d, ["", "", "", ""])
        damages.append({"name": d, "materials": mats, "severity": sev})
    return {"components": comps, "damages": damages, "causesByDamage": edif_causes}


# ───────────── PUENTE (de Gravedades_Base.xlsx, enriquecido) ─────────────
def clean(s):
    return "" if s is None else str(s).strip()


def build_puente():
    wb = openpyxl.load_workbook("Gravedades_Base.xlsx", data_only=True)
    # componentes -> elementos
    rows = list(wb["Componentes"].iter_rows(values_only=True))
    comp_names = [clean(rows[1][c]) for c in range(6)]
    # materiales y zonas típicas por elemento de puente (enriquecido, opción A)
    br_mat = {
        "Fundación": [HA], "Muro frontal portante": [HA], "Guardalastres": [HA], "Alas": [HA],
        "Aparato de apoyo": ["Elastomérico", AC], "Cama de nivelación": [HA],
        "Columnas": [HA, AC], "Línea de contención": [HA], "Longuerina": [HA, HP, AC],
        "Cabezal superior": [HA], "Cabezal inferior": [HA], "Diagonal": [AC, HA], "Montante": [AC, HA],
        "Voladizo": [HA, HP], "Travesaño intermedia": [HA, HP], "Travesaño Apoyo": [HA, HP],
        "Contraventación": [AC], "Vigas": [HA, HP, AC], "Losa": [HA, HP], "Cartela": [HA, AC],
    }
    _wall = ["Base", "Cuerpo/paño", "Coronación", "Encuentro", "Cara vista", "Junta", "Otro"]
    _cap = ["Apoyo", "Centro", "Voladizo", "Cara inferior", "Nudo", "Otro"]
    _travesano = ["Apoyo", "Centro de vano", "Fondo", "Anclaje de pretensado", "Otro"]
    br_zone = {
        "default": ["Base", "Tercio central", "Cabeza/extremo", "Cara vista", "Otro"],
        "Columnas": ["Base", "Anclaje / placa base", "Fuste/tercio central", "Cabeza/capitel", "Nudo", "Otro"],
        "Vigas": ["Apoyo/extremo", "Centro de vano", "Fondo de viga", "Anclaje de pretensado", "Cara lateral", "Otro"],
        "Longuerina": ["Apoyo", "Centro de vano", "Anclaje de pretensado", "Otro"],
        "Losa": ["Centro de paño", "Borde/perímetro", "Cara inferior", "Junta", "Otro"],
        "Fundación": ["Cara superior", "Cara lateral", "Descalce", "Otro"],
        "Aparato de apoyo": ["Cuerpo", "Anclaje", "Otro"],
        "Diagonal": ["Tramo", "Nudo/conexión", "Anclaje", "Otro"],
        "Contraventación": ["Tramo", "Conexión", "Anclaje", "Otro"],
        # muros y elementos tipo pantalla
        "Muro frontal portante": _wall,
        "Guardalastres": ["Coronación", "Cuerpo", "Encuentro con losa", "Cara vista", "Otro"],
        "Alas": _wall,
        "Línea de contención": ["Tramo", "Anclaje", "Cara vista", "Junta", "Otro"],
        "Cama de nivelación": ["Cuerpo", "Contacto con apoyo", "Otro"],
        # cabezales (pier cap / cordones)
        "Cabezal superior": _cap,
        "Cabezal inferior": _cap,
        # travesaños, voladizo, cartela, montante
        "Travesaño intermedia": _travesano,
        "Travesaño Apoyo": _travesano,
        "Voladizo": ["Empotramiento", "Extremo", "Cara inferior", "Anclaje de pretensado", "Otro"],
        "Cartela": ["Nudo", "Cara vista", "Conexión", "Otro"],
        "Montante": ["Tramo", "Nudo/conexión", "Anclaje", "Otro"],
    }
    comps = []
    for c in range(6):
        name = clean(rows[1][c])
        if not name:
            continue
        elems = []
        seen = set()
        for r in range(2, len(rows)):
            e = clean(rows[r][c])
            # descarta vacíos, duplicados y filas "nombre_del_componente" (con _)
            if not e or e == name or e.replace("_", " ") == name or e in seen:
                continue
            seen.add(e)
            elems.append({"element": e, "materials": br_mat.get(e, [HA]),
                          "zones": br_zone.get(e, br_zone["default"])})
        if elems:
            comps.append({"component": name, "elements": elems})
    # anclaje como componente propio (opción A)
    comps.append({"component": "Anclajes y conexiones", "elements": [
        {"element": "Anclaje de pretensado", "materials": [HP, AC], "zones": ["Zona de anclaje", "Placa", "Otro"]},
        {"element": "Anclaje / placa base", "materials": [AC], "zones": ["Placa base", "Perno de anclaje", "Otro"]},
    ]})
    # deteriorios + severity
    damages = []
    for r in wb["Gravedad"].iter_rows(min_row=2, values_only=True):
        d = clean(r[1])
        if not d:
            continue
        sev = [clean(r[2]), clean(r[3]), clean(r[4]), clean(r[5])]
        damages.append({"name": d, "materials": [TODOS], "severity": sev})
    # causas por deterioro
    crows = list(wb["Causas"].iter_rows(values_only=True))
    header = [clean(x) for x in crows[0]]
    causes = {}
    for c in range(1, len(header)):
        d = header[c]
        if not d:
            continue
        cs = [clean(crows[r][c]) for r in range(2, len(crows)) if clean(crows[r][c])]
        if cs:
            causes[d] = cs
    return {"components": comps, "damages": damages, "causesByDamage": causes}


catalog = {"edificio": build_edificio(), "puente": build_puente()}

ts = []
ts.append("// GENERADO por scripts/build_catalog_ts.py — no editar a mano.")
ts.append("// Catálogo de inspección por tipo de estructura (edificio, puente).")
ts.append("")
ts.append("export interface CatalogElement { element: string; materials: string[]; zones: string[] }")
ts.append("export interface CatalogComponent { component: string; elements: CatalogElement[] }")
ts.append("export interface CatalogDamage { name: string; materials: string[]; severity: string[] }")
ts.append("export interface StructureCatalog {")
ts.append("  components: CatalogComponent[]")
ts.append("  damages: CatalogDamage[]")
ts.append("  causesByDamage: Record<string, string[]>")
ts.append("}")
ts.append("")
ts.append("export const MATERIALS: string[] = " + json.dumps(MATERIALS, ensure_ascii=False))
ts.append("")
ts.append("export const CATALOG: Record<string, StructureCatalog> = " + json.dumps(catalog, ensure_ascii=False, indent=2))
ts.append("")

with open("src/data/catalog.ts", "w", encoding="utf-8") as f:
    f.write("\n".join(ts))

print("OK src/data/catalog.ts")
for k, v in catalog.items():
    print(f"  {k}: {len(v['components'])} componentes, "
          f"{sum(len(c['elements']) for c in v['components'])} elementos, "
          f"{len(v['damages'])} deterioros, {len(v['causesByDamage'])} causas")
