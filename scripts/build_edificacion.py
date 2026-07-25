# -*- coding: utf-8 -*-
"""Genera Edificacion_Base_v2.xlsx (análogo a Gravedades_Base.xlsx) con materialidad y anclajes."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
hdr = Font(bold=True, color="FFFFFF")
fill = PatternFill("solid", fgColor="1E293B")


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)


# ── 0) MATERIALES (catálogo) ──
materiales = [
    ("Hormigón armado", "HA"),
    ("Hormigón pretensado/postensado", "HP"),
    ("Acero", "Acero"),
    ("Madera", "Madera"),
    ("Albañilería confinada", "Alb. confinada"),
    ("Albañilería armada", "Alb. armada"),
    ("Albañilería simple", "Alb. simple"),
    ("Mixto acero-hormigón", "Mixto"),
    ("Otro", "Otro"),
]
wsm = wb.active
wsm.title = "Materiales"
wsm.append(["Material", "Abrev."])
for m, a in materiales:
    wsm.append([m, a])
style_header(wsm, 2)
wsm.column_dimensions["A"].width = 34
wsm.column_dimensions["B"].width = 16

HA = "Hormigón armado"
HP = "Hormigón pretensado/postensado"
AC = "Acero"
MA = "Madera"
ALBC = "Albañilería confinada"
ALBA = "Albañilería armada"
MX = "Mixto acero-hormigón"

# ── 1) COMPONENTES -> ELEMENTOS (con materiales aplicables y zonas) ──
# (elemento, [materiales], "zonas; ...")
componentes = {
    "Fundaciones": [
        ("Zapata aislada", [HA], "Cara superior; Pedestal; Perímetro; Otro"),
        ("Zapata corrida", [HA], "Cara superior; Cara lateral; Otro"),
        ("Viga de fundación", [HA], "Apoyo; Centro de vano; Nudo; Otro"),
        ("Losa de fundación (radier)", [HA], "Centro de paño; Borde; Junta; Otro"),
        ("Pilote", [HA, AC, MA], "Cabeza; Fuste; Punta; Anclaje/empotramiento; Otro"),
        ("Cabezal de pilotes", [HA], "Cara superior; Encuentro con pilote; Otro"),
        ("Sobrecimiento", [HA, ALBC], "Base; Coronación; Otro"),
        ("Muro de subterráneo", [HA], "Base; Coronación; Paño central; Junta; Otro"),
    ],
    "Muros estructurales": [
        ("Muro estructural / de corte", [HA], "Base; Coronación; Paño central; Encuentro/esquina; Vano/abertura; Otro"),
        ("Machón", [HA, ALBC], "Base; Cuerpo; Encuentro; Otro"),
        ("Muro de albañilería confinada", [ALBC], "Paño central; Encuentro; Base; Bajo cadena/pilar; Vano; Otro"),
        ("Muro de albañilería armada", [ALBA], "Paño central; Encuentro; Base; Vano; Otro"),
    ],
    "Pilares y columnas": [
        ("Columna / Pilar", [HA, AC, MA, MX], "Base; Anclaje / placa base; Tercio central; Cabeza/capitel; Nudo; Empalme; Otro"),
    ],
    "Vigas y cadenas": [
        ("Viga", [HA, HP, AC, MA, MX], "Apoyo/extremo; Centro de vano; Fondo de viga; Nudo; Anclaje de pretensado; Anclaje de conexión; Cara lateral; Otro"),
        ("Cadena", [HA], "Tramo; Encuentro; Otro"),
        ("Dintel", [HA, AC, ALBC], "Centro; Apoyo; Otro"),
    ],
    "Losas y entrepisos": [
        ("Losa", [HA, HP], "Centro de paño; Borde/perímetro; Cara inferior; Cara superior; Junta; Otro"),
        ("Losa colaborante", [MX], "Centro de paño; Apoyo; Conexión de corte; Otro"),
        ("Vigueta", [HA, HP, AC, MA], "Apoyo; Centro de vano; Fondo; Otro"),
        ("Voladizo / balcón", [HA], "Empotramiento; Extremo; Cara inferior; Otro"),
    ],
    "Nudos y conexiones": [
        ("Nudo viga-columna", [HA, AC], "Núcleo del nudo; Extremo de viga; Extremo de columna; Otro"),
        ("Conexión soldada", [AC], "Cordón de soldadura; Placa; Otro"),
        ("Conexión apernada", [AC], "Perno; Placa; Otro"),
        ("Anclaje", [AC, HA], "Placa base; Perno de anclaje; Anclaje de pretensado; Anclaje químico; Anclaje mecánico; Otro"),
        ("Empalme de armadura", [HA], "Traslapo; Manguito; Otro"),
        ("Apoyo / junta de dilatación", [HA, AC, "Elastomérico"], "Aparato de apoyo; Junta; Otro"),
    ],
    "Escaleras": [
        ("Losa de escalera", [HA], "Tramo; Descanso; Apoyo; Otro"),
        ("Peldañeado", [HA, MA, AC], "Peldaño; Nariz; Otro"),
        ("Zanca / viga de escalera", [HA, AC, MA], "Apoyo; Tramo; Anclaje; Otro"),
        ("Rampa", [HA], "Tramo; Apoyo; Otro"),
    ],
    "Techumbre estructural": [
        ("Cercha", [AC, MA], "Cordón superior; Cordón inferior; Diagonal; Montante; Nudo; Anclaje/apoyo; Otro"),
        ("Viga de techo", [AC, MA, HA], "Apoyo; Centro de vano; Anclaje; Otro"),
        ("Costanera / correa", [AC, MA], "Apoyo; Tramo; Conexión; Otro"),
        ("Losa de cubierta", [HA], "Centro de paño; Borde; Junta; Otro"),
    ],
    "Tabiquería": [
        ("Tabique", [ALBC, "Tabiquería seca"], "Paño central; Encuentro; Base; Vano; Otro"),
        ("Antepecho / parapeto", [HA, ALBC], "Coronación; Base; Otro"),
    ],
}
ws = wb.create_sheet("Componentes")
ws.append(["Componente", "Elemento", "Materiales aplicables", "Zonas típicas (sugeridas)"])
for comp, elems in componentes.items():
    for name, mats, zonas in elems:
        ws.append([comp, name, "; ".join(mats), zonas])
style_header(ws, 4)
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 46
ws.column_dimensions["D"].width = 62

# ── 2) DETERIOROS: materiales aplicables + criterios de gravedad ──
TODOS = "todos"
# deterioro: ([materiales], min, media, alta, muy alta)
det = {
    "Fisuras": ([HA, "Albañilería", MA], "abertura < 0,2 mm, sin incidencia estructural", "0,2-0,4 mm o incidencia estructural leve", "0,4-1 mm con incidencia estructural", "> 1 mm, evolutiva o generalizada"),
    "Grietas": ([HA, "Albañilería"], "-", "1-2 mm con incidencia estructural", "2-5 mm", "> 5 mm, evolutiva; compromete estabilidad"),
    "Fisuras en mapa o retícula": ([HA], "fisuras finas < 0,4 mm", "0,4-2 mm", "> 2 mm", "generalizada con disgregación"),
    "Descascaramiento sin armadura a la vista": ([HA], "profundidad < 10 mm", "10-100 mm", "> 100 mm", "pérdida de sección relevante"),
    "Descascaramiento con armadura a la vista": ([HA], "sin pérdida de sección de armadura", "pérdida de sección leve (<10%)", "pérdida apreciable (10-50%)", "pérdida > 50% o barras seccionadas"),
    "Armadura a la vista": ([HA], "puntual, sin pérdida de sección", "pérdida de sección leve", "pérdida apreciable", "pérdida severa / barras cortadas"),
    "Corrosión de armaduras/acero": ([HA, AC], "oxidación superficial, <10% sup., sin pérdida", "pérdida de sección leve (10-25%)", "pérdida apreciable (25-50%)", "pérdida > 50% o seccionamiento"),
    "Mancha de óxido": ([HA, AC], "mancha superficial aislada", "manchas extendidas", "asociada a fisura/desconchón", "generalizada con corrosión activa"),
    "Coqueras / nidos de grava": ([HA], "afecta < 50% del paramento", "> 50% del paramento", "con armadura expuesta", "compromete sección resistente"),
    "Humedades / filtraciones": ([TODOS], "mancha superficial, sin daño", "humedad persistente, inicio de deterioro", "filtración activa con daño", "filtración severa con corrosión/pérdida"),
    "Eflorescencias": ([HA, "Albañilería"], "depósitos leves aislados", "extendidas", "con disgregación superficial", "generalizada con pérdida de material"),
    "Deformación / flecha excesiva": ([TODOS], "perceptible, < L/500", "L/500-L/250", "L/250-L/150", "> L/150 o creciente"),
    "Falta de alineación / desplome": ([TODOS], "leve, solo estético", "perceptible, funcional", "importante, riesgo", "severo, inestabilidad"),
    "Disgregación / pulverización": ([HA, "Albañilería"], "superficial, sin pérdida de sección", "pérdida leve de sección", "pérdida apreciable", "pérdida severa"),
    "Aplastamiento (albañilería)": (["Albañilería", HA], "fisuración local", "aplastamiento parcial", "aplastamiento con pérdida de unidades", "colapso local"),
    "Pandeo local": ([AC], "deformación < 1 mm", "1-5 mm", "> 5 mm", "pandeo con pérdida de capacidad"),
    "Pérdida de mortero de junta": (["Albañilería"], "puntual, superficial", "junta abierta local", "generalizada", "compromete estabilidad del muro"),
    "Pérdida de tratamiento protector": ([AC, MA], "puntual, estético", "extendido", "con inicio de corrosión/pudrición", "generalizado con daño"),
    "Pudrición / xilófagos": ([MA], "superficial", "pérdida de sección leve", "pérdida apreciable", "pérdida severa / colapso"),
    "Deterioro de soldadura / conexión": ([AC], "sin pérdida de capacidad", "fisura leve", "fisura importante", "rotura de conexión"),
    "Pérdida / aflojamiento de pernos y anclajes": ([AC, HA], "aflojamiento puntual", "falta de algún perno/anclaje", "falta de varios", "conexión/anclaje inutilizado"),
    "Alteración superficial": ([TODOS], "muy superficial, sin pérdida de sección", "ligera pérdida de recubrimiento", "pérdida apreciable", "pérdida severa"),
    "Vegetación / biológico": ([TODOS], "no leñosa, sin daño", "con inicio de daño", "leñosa con daño", "daño estructural"),
}
ws2 = wb.create_sheet("Gravedad")
ws2.append(["Deterioro", "Materiales", "mínima", "media", "alta", "muy alta"])
for d, (mats, a, b, c, e) in det.items():
    ws2.append([d, "; ".join(mats), a, b, c, e])
style_header(ws2, 6)
ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 24
for col in "CDEF":
    ws2.column_dimensions[col].width = 32

# ── 3) CAUSAS por deterioro ──
causas = {
    "Fisuras": ["retracción del hormigón", "dilatación/contracción térmica", "curado inadecuado", "asiento diferencial", "sobrecarga", "deficiente ejecución", "causa desconocida"],
    "Grietas": ["acción sísmica", "asiento diferencial", "sobrecarga gravitacional", "esfuerzos (corte)", "esfuerzos (flexión)", "corrosión de armaduras", "empuje de terreno", "deficiente diseño", "causa desconocida"],
    "Fisuras en mapa o retícula": ["retracción", "ataque químico", "reacción árido-álcali", "corrosión de armaduras", "ciclos hielo-deshielo", "deficiente ejecución"],
    "Descascaramiento sin armadura a la vista": ["golpe o impacto", "ciclos hielo-deshielo", "ataque químico", "deficiente ejecución", "causa desconocida"],
    "Descascaramiento con armadura a la vista": ["corrosión de armaduras", "carbonatación", "ataque de cloruros", "escasez de recubrimiento", "deficiente ejecución"],
    "Armadura a la vista": ["escasez de recubrimiento", "corrosión de armaduras", "desconchón previo", "deficiente ejecución"],
    "Corrosión de armaduras/acero": ["carbonatación", "ataque de cloruros", "humedad/filtración", "escasez de recubrimiento", "ambiente agresivo/marino", "falta de mantención"],
    "Mancha de óxido": ["corrosión de armaduras", "humedad", "escorrentía superficial", "falta de limpieza del encofrado"],
    "Coqueras / nidos de grava": ["deficiente ejecución (vibrado/hormigonado)", "deficiente dosificación", "causa desconocida"],
    "Humedades / filtraciones": ["impermeabilización defectuosa", "rotura de conducción", "capilaridad", "condensación", "mal drenaje", "falta de mantención"],
    "Eflorescencias": ["humedad/filtración", "capilaridad", "impermeabilización defectuosa", "sales del material"],
    "Deformación / flecha excesiva": ["sobrecarga", "infradimensionamiento", "fluencia (creep) del hormigón", "cambio de uso", "esfuerzos (flexión)", "pandeo"],
    "Falta de alineación / desplome": ["asiento diferencial", "acción sísmica", "empuje de terreno", "deficiente ejecución"],
    "Disgregación / pulverización": ["ataque químico", "ciclos hielo-deshielo", "carbonatación avanzada", "curado inadecuado", "calidad deficiente del hormigón"],
    "Aplastamiento (albañilería)": ["sobrecarga", "acción sísmica", "infradimensionamiento", "concentración de esfuerzos"],
    "Pandeo local": ["sobrecarga (compresión)", "esbeltez excesiva", "arriostramiento insuficiente", "infradimensionamiento", "impacto"],
    "Pérdida de mortero de junta": ["envejecimiento", "humedad", "ataque químico", "deficiente ejecución", "acción sísmica"],
    "Pérdida de tratamiento protector": ["envejecimiento", "acción climática", "falta de mantención", "corrosión", "ataque químico"],
    "Pudrición / xilófagos": ["humedad", "hongos", "insectos xilófagos", "falta de protección/mantención"],
    "Deterioro de soldadura / conexión": ["fatiga", "corrosión", "deficiente ejecución", "esfuerzos cíclicos", "sobrecarga"],
    "Pérdida / aflojamiento de pernos y anclajes": ["corrosión", "aflojamiento por vibración", "deficiente ejecución", "esfuerzos cíclicos", "sobrecarga", "acción sísmica"],
    "Alteración superficial": ["acción climática", "envejecimiento", "ataque químico", "abrasión", "pérdida de tratamiento"],
    "Vegetación / biológico": ["humedad", "sedimentación orgánica", "falta de mantención"],
}
ws3 = wb.create_sheet("Causas")
ws3.append(["Deterioro", "Causas probables"])
for d, cs in causas.items():
    ws3.append([d, "; ".join(cs)])
style_header(ws3, 2)
ws3.column_dimensions["A"].width = 36
ws3.column_dimensions["B"].width = 95

wb.save("Edificacion_Base_v2.xlsx")
print("OK Edificacion_Base_v2.xlsx")
print("Materiales:", len(materiales))
print("Componentes:", len(componentes), "| Elementos:", sum(len(v) for v in componentes.values()))
print("Deterioros:", len(det), "| Causas:", len(causas))
